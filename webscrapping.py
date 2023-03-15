import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import locale
from locale import atof 

# clean up txt file before executing script
open('stock_ticket_errors.txt', 'w').close() 

stocksFields = {
		'stock_ticket': [],
        'ValorPatri': [],
        'ValorCota': [],
        'DyAno': [],
        'UltMesPercentual': [],
        'UltMesAbsoluto': []
}

# variable to store all column names from excel file
columnName = {}

# my user agent. To make multiple requests without getting blocked
headers= {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:107.0) Gecko/20100101 Firefox/107.0'
}

index = 0
listStocks = []

# Phase 1: Open excel
spreadsheet = load_workbook('ControleFII.xlsx')
activeWorksheet = spreadsheet.active

# Phase 1.1: getting last empty row...
lastRowWithContent = len(list(activeWorksheet.rows))-1

print('Creates a dict of column names...')

# Phase 2: Extract Column Names from excel file
for COLUMN in activeWorksheet.iter_cols(1, activeWorksheet.max_column):
	columnName[COLUMN[0].value] = index
	index += 1
# Phase 2.1: Extract all stocks names from activeWorksheet
for row_cells in activeWorksheet.iter_rows(min_row=1, max_row=lastRowWithContent):
    listStocks.append(row_cells[columnName['Fundo']].value)
# remove row 1 because is the name of the Column, and remove none values
listStocks.pop(0)
listStocks = [i for i in listStocks if i is not None]

print(f'Amount of stocks to watch: {len(listStocks)}')
print(f'Making requests to statusinvest...')

i = 0

# Phase 3: Loop through the list of stocks and populate dictionary of stocks
for i in range(0, len(listStocks)):

    # make url requests
    url = f'https://statusinvest.com.br/fundos-imobiliarios/{listStocks[i]}'
    site = requests.get(url, headers=headers)
    
    print(f'fazendo requisição..: {url}')
    
    # parsing document using beatifulsoup4...'
    soup = BeautifulSoup(site.content, 'html.parser')
    
    # match from site desired content to extract
    strongClassArray = soup.find_all('strong', class_ = "value")[0:17]
    
    # if request contains desired content extract them
    if strongClassArray != None and len(strongClassArray) > 3:
        DyAno = soup.find('span', string = "%").findPrevious('strong').getText().strip()
        ValorPatri = soup.find('h3', string = "Val. patrimonial p/cota").findNext('strong').getText().strip()
        ValorCotaArray= soup.find('small', string = "Cotação base").find_all_next('b', class_ = 'sub-value')
        ValorCota = ValorCotaArray[4].getText().strip() if ValorCotaArray[4].getText().strip() != '-' else ValorCotaArray[0].getText().strip()
        UltMesPercentualArray = soup.find('small', string = "Rendimento").find_all_next('b', class_ = 'sub-value')
        UltMesPercentual = UltMesPercentualArray[4].getText().strip() if UltMesPercentualArray[4].getText().strip() != '-' else UltMesPercentualArray[0].getText().strip()
        UltMesAbsolutoArray = soup.find('span', string = "Último rendimento").find_all_next('strong', class_ = 'value')[0:2]
        UltMesAbsoluto = UltMesAbsolutoArray[1].getText().strip() if UltMesAbsolutoArray[1].getText().strip() != '-' else UltMesAbsolutoArray[0].getText().strip()
    
        print(f' DyAno: {DyAno}')
        print(f' ValorPatri: {ValorPatri}')
        print(f' ValorCota: {ValorCota}')
        print(f' UltMesPercentual: {UltMesPercentual}')
        print(f' UltMesAbsoluto: {UltMesAbsoluto}')

        stocksFields['DyAno'].append(DyAno)
        stocksFields['ValorPatri'].append(ValorPatri)
        stocksFields['ValorCota'].append(ValorCota)
        stocksFields['UltMesPercentual'].append(UltMesPercentual)
        stocksFields['UltMesAbsoluto'].append(UltMesAbsoluto)
        stocksFields['stock_ticket'].append(listStocks[i])
        
    # if not, fill values with 0
    else :
        stocksFields['DyAno'].append("0")
        stocksFields['ValorPatri'].append("0")
        stocksFields['ValorCota'].append("0")
        stocksFields['UltMesPercentual'].append("0")
        stocksFields['UltMesAbsoluto'].append("0")
        stocksFields['stock_ticket'].append(listStocks[i])
        # write to file, stock tickets with error. Note: they won't be updated on excel file
        file = open("stock_ticket_errors.txt", "a")
        file.write(listStocks[i] + "\n")
        

# check length of all values, to not get an error when transforming to dataframe       
print('len DyAno',len(stocksFields['DyAno']))
print('len ValorPatri',len(stocksFields['ValorPatri']))
print('len ValorCota',len(stocksFields['ValorCota']))
print('len stock_ticket',len(stocksFields['stock_ticket']))
        
dataFrame = pd.DataFrame(stocksFields) #use dictionary to create a dataframe
print(dataFrame.tail())

print(f'Replacing invalid values...')
# If values are a string, replace them with zero. To no get a error when converting to float
dataFrame.loc[dataFrame.ValorCota == '-', 'ValorCota'] = "0"
dataFrame.loc[dataFrame.UltMesPercentual == '-', 'UltMesPercentual'] = "0"
dataFrame.loc[dataFrame.UltMesAbsoluto == '-', 'UltMesAbsoluto'] = "0"

#Converting values to float...
locale.setlocale(locale.LC_NUMERIC, '')
dataFrame['ValorPatri'] = dataFrame['ValorPatri'].map(atof)
dataFrame['ValorCota'] = dataFrame['ValorCota'].map(atof)
dataFrame['DyAno'] = dataFrame['DyAno'].map(atof)
dataFrame['UltMesPercentual'] = dataFrame['UltMesPercentual'].map(atof)
dataFrame['UltMesAbsoluto'] = dataFrame['UltMesAbsoluto'].map(atof) 
                   
i = 0

# Phase 4: Populate excel file with values from dataframe
for row_cells in activeWorksheet.iter_rows(min_row=2, max_row=len(stocksFields['stock_ticket'])+1):

    # determine column letter in ASCII, and position of row. Eg: A3, A4, A5, B1, B2 ...
    
    letraColunaNomeFundo = row_cells[columnName['Fundo']].column + 64
    linhaNomefundo = row_cells[columnName['Fundo']].row
    
    
    letraColunaPatri = row_cells[columnName['ValorPatri']].column + 64
    linhaPatri = row_cells[columnName['ValorPatri']].row
     
    letraColunaValorCota = row_cells[columnName['ValorCota']].column + 64
    linhaValorCota = row_cells[columnName['ValorCota']].row
    
    letraColunaDy = row_cells[columnName['Ano']].column + 64
    linhaDy= row_cells[columnName['Ano']].row
    
    letraColunaUltMesPercentual = row_cells[columnName['Últ.Mês%']].column + 64
    linhaUltMesPercentual = row_cells[columnName['Últ.Mês%']].row

    letraColunaUltMesAbsoluto = row_cells[columnName['Últ.Mês.Abs']].column + 64
    linhaUltMesAbsoluto = row_cells[columnName['Últ.Mês.Abs']].row
    

    #update only with column "Fundo" from excel, matches column stock_ticket from dataframe
    if (row_cells[columnName['Fundo']].value) == dataFrame['stock_ticket'][i]:
        
        # appedning values to corresponding column and row cells
        if dataFrame['ValorPatri'][i] != 0:
            activeWorksheet[chr(letraColunaPatri) + str(linhaPatri)] = dataFrame['ValorPatri'][i]
        if dataFrame['DyAno'][i] != 0:
            activeWorksheet[chr(letraColunaDy) + str(linhaDy)] = dataFrame['DyAno'][i]
        if dataFrame['UltMesAbsoluto'][i] != 0:
            activeWorksheet[chr(letraColunaUltMesAbsoluto) + str(linhaUltMesAbsoluto)] = dataFrame['UltMesAbsoluto'][i]
        if dataFrame['UltMesPercentual'][i] != 0:
            activeWorksheet[chr(letraColunaUltMesPercentual) + str(linhaUltMesPercentual)] = dataFrame['UltMesPercentual'][i]
        if dataFrame['ValorCota'][i] != 0:
            activeWorksheet[chr(letraColunaValorCota) + str(linhaValorCota)] = dataFrame['ValorCota'][i]
            
    # increment variable only if line is not empty
    if type(row_cells[columnName['Fundo']].value) == str:
        i = i + 1
    
spreadsheet.save('ControleFII.xlsx')
print(dataFrame.tail())

# TODO: escrever função para colocar nome dos ativos que não foram atualizados em algum txt para checagem posterior...